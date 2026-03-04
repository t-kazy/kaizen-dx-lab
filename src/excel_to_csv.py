"""Excel/CSVシフトデータをCSV形式に自動変換するツール。

典型的な日本語シフト表（行=従業員、列=日付）を読み取り、
「日付, 従業員名, シフト種別」のフラットなCSVに変換する。
Excel (.xlsx) と CSV (.csv) の両方を入力としてサポートする。
"""

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


class _CsvCell:
    """openpyxl の Cell を模倣する軽量オブジェクト。"""

    def __init__(self, value):
        self.value = value


class CsvWorksheet:
    """CSV ファイルを openpyxl ワークシート風に扱うラッパー。

    既存の detect_header_row / detect_name_column / parse_dates /
    find_data_rows / convert_sheet をそのまま再利用できる。
    """

    def __init__(self, rows, title="CSV"):
        self.title = title
        self._rows = rows  # list[list[str|None]]
        self.max_row = len(rows)
        self.max_column = max((len(r) for r in rows), default=0)

    def cell(self, row, column):
        r = row - 1
        c = column - 1
        if 0 <= r < len(self._rows) and 0 <= c < len(self._rows[r]):
            return _CsvCell(self._rows[r][c] or None)
        return _CsvCell(None)

    @classmethod
    def from_file(cls, csv_path, encoding="utf-8-sig"):
        """CSVファイルを読み込んで CsvWorksheet を返す。"""
        rows = []
        path = Path(csv_path)
        # utf-8-sig → utf-8 → shift_jis → cp932 の順で試行
        encodings = [encoding, "utf-8", "shift_jis", "cp932"]
        for enc in encodings:
            try:
                with open(path, newline="", encoding=enc) as f:
                    reader = csv.reader(f)
                    rows = [row for row in reader]
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        else:
            raise ValueError(f"CSVファイルのエンコーディングを検出できません: {path}")
        return cls(rows, title=path.stem)


def convert_csv_file(csv_path, year=None, encoding="utf-8-sig"):
    """CSVシフト表ファイルを変換する。

    テーブル形式（1行1レコード、ヘッダーに「日付」「シフト」等がある）を
    自動検出し、該当すればそちらで処理する。
    検出できなければ従来のマトリクス形式として処理する。
    """
    ws = CsvWorksheet.from_file(csv_path, encoding=encoding)

    # テーブル形式かどうかを判定
    tabular = _detect_tabular_csv(ws)
    if tabular is not None:
        return tabular

    return convert_sheet(ws, year=year)


# テーブル形式CSVのヘッダー検出用キーワード
_DATE_KEYWORDS = ("日付", "日にち", "勤務日", "出勤日", "date")
_EMPLOYEE_KEYWORDS = ("指導員名", "従業員名", "スタッフ名", "担当者", "氏名", "名前", "employee", "name")
_SHIFT_KEYWORDS = ("シフト名", "シフト種別", "シフト", "勤務種別", "勤務区分", "shift")


def _detect_tabular_csv(ws):
    """テーブル形式CSV（1行1レコード）を検出して変換する。

    ヘッダー行に「日付」「従業員名/指導員名」「シフト名」に相当する列が
    あればテーブル形式と判定する。検出できなければ None を返す。
    """
    if ws.max_row < 2:
        return None

    # 1行目をヘッダーとして検査
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        headers.append(str(val).strip() if val else "")

    headers_lower = [h.lower() for h in headers]

    date_col = _find_column(headers_lower, _DATE_KEYWORDS)
    employee_col = _find_column(headers_lower, _EMPLOYEE_KEYWORDS)
    shift_col = _find_column(headers_lower, _SHIFT_KEYWORDS)

    if date_col is None or employee_col is None or shift_col is None:
        return None

    records = []
    for row_idx in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=date_col + 1).value
        employee_val = ws.cell(row=row_idx, column=employee_col + 1).value
        shift_val = ws.cell(row=row_idx, column=shift_col + 1).value

        if not date_val or not employee_val or not shift_val:
            continue

        date_str = _normalize_date(date_val, datetime.now().year)
        if not date_str:
            date_str = str(date_val).strip()

        records.append({
            "date": date_str,
            "employee": str(employee_val).strip(),
            "shift": str(shift_val).strip(),
        })

    return records


def _find_column(headers, keywords):
    """ヘッダーリストからキーワードに一致する列のインデックスを返す。"""
    for i, h in enumerate(headers):
        for kw in keywords:
            if kw in h:
                return i
    return None


# --- freee勤怠管理Plus 変換 ---

_START_TIME_KEYWORDS = ("勤務時間(開始)", "勤務時間（開始）", "開始時刻", "出勤時刻")
_END_TIME_KEYWORDS = ("勤務時間(終了)", "勤務時間（終了）", "終了時刻", "退勤時刻")


def load_freee_mapping(mapping_path):
    """freeeマッピング設定ファイルを読み込む。"""
    path = Path(mapping_path)
    if not path.exists():
        raise FileNotFoundError(f"マッピングファイルが見つかりません: {path}")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def convert_to_freee(csv_path, mapping_path, encoding="utf-8-sig"):
    """CSVシフト表をfreee勤怠管理Plus スケジュールCSV形式に変換する。"""
    ws = CsvWorksheet.from_file(csv_path, encoding=encoding)
    mapping = load_freee_mapping(mapping_path)

    employee_map = mapping.get("employee_map", {})
    shift_map = mapping.get("shift_map", {})
    time_to_pattern = mapping.get("time_to_pattern", {})
    skip_shifts = set(mapping.get("skip_shifts", []))

    # ヘッダー検出
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        headers.append(str(val).strip() if val else "")
    headers_lower = [h.lower() for h in headers]

    date_col = _find_column(headers_lower, _DATE_KEYWORDS)
    employee_col = _find_column(headers_lower, _EMPLOYEE_KEYWORDS)
    shift_col = _find_column(headers_lower, _SHIFT_KEYWORDS)
    start_time_col = _find_column(headers, _START_TIME_KEYWORDS)
    end_time_col = _find_column(headers, _END_TIME_KEYWORDS)

    if date_col is None or employee_col is None:
        raise ValueError("CSVから日付列または従業員列を検出できませんでした。")

    records = []
    unmapped_employees = set()
    unmapped_shifts = set()

    for row_idx in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=date_col + 1).value
        employee_val = ws.cell(row=row_idx, column=employee_col + 1).value

        if not date_val or not employee_val:
            continue

        shift_val = ws.cell(row=row_idx, column=shift_col + 1).value if shift_col is not None else None
        shift_str = str(shift_val).strip() if shift_val else ""

        if shift_str in skip_shifts or not shift_str:
            continue

        # 従業員名 → コード
        emp_name = str(employee_val).strip()
        emp_code = _resolve_employee_code(emp_name, employee_map)
        if not emp_code:
            unmapped_employees.add(emp_name)
            continue

        # シフト → パターンコード
        pattern_code = shift_map.get(shift_str)
        if not pattern_code and start_time_col is not None and end_time_col is not None:
            start_val = ws.cell(row=row_idx, column=start_time_col + 1).value
            end_val = ws.cell(row=row_idx, column=end_time_col + 1).value
            if start_val and end_val:
                time_key = f"{_normalize_time(start_val)}-{_normalize_time(end_val)}"
                pattern_code = time_to_pattern.get(time_key)
        if not pattern_code:
            unmapped_shifts.add(shift_str)
            continue

        # 日付正規化 (freee は YYYY/MM/DD 形式)
        date_str = _normalize_date(date_val, datetime.now().year)
        if date_str:
            date_str = date_str.replace("-", "/")
        else:
            date_str = str(date_val).strip()

        records.append({
            "勤務日": date_str,
            "従業員コード": emp_code,
            "パターンコード": pattern_code,
        })

    for name in sorted(unmapped_employees):
        print(f"警告: 従業員 '{name}' のコードが未設定です (スキップ)", file=sys.stderr)
    for shift in sorted(unmapped_shifts):
        print(f"警告: シフト '{shift}' のパターンコードが未設定です (スキップ)", file=sys.stderr)

    return records


def _resolve_employee_code(name, employee_map):
    """従業員名からコードを解決する（スペース有無を吸収）。"""
    if name in employee_map:
        return employee_map[name]
    name_no_space = name.replace(" ", "").replace("\u3000", "")
    for key, code in employee_map.items():
        if key.replace(" ", "").replace("\u3000", "") == name_no_space:
            return code
    return None


def _normalize_time(val):
    """時刻を H:MM 形式に正規化する。"""
    s = str(val).strip().replace("：", ":")
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            h = int(parts[0])
            m = int(parts[1])
            return f"{h}:{m:02d}"
        except ValueError:
            pass
    return s


def write_freee_csv(records, output_path, encoding="cp932"):
    """freee勤怠管理Plus形式のCSVを書き出す。ヘッダーなし・Shift_JIS。"""
    with open(output_path, "w", newline="", encoding=encoding) as f:
        writer = csv.writer(f)
        for record in sorted(records, key=lambda r: (r["勤務日"], r["従業員コード"])):
            writer.writerow([record["勤務日"], record["従業員コード"], record["パターンコード"]])
    return len(records)


def detect_header_row(ws, max_scan=20):
    """日付ヘッダー行を自動検出する。

    数値・日付セルが連続して並ぶ行をヘッダーとみなす。
    """
    best_row = None
    best_count = 0

    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        date_count = 0
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if _is_date_like(cell.value):
                date_count += 1
        if date_count > best_count:
            best_count = date_count
            best_row = row_idx

    if best_count < 2:
        return None
    return best_row


def _is_date_like(value):
    """値が日付らしいかどうかを判定する。"""
    if value is None:
        return False
    if isinstance(value, datetime):
        return True
    s = str(value).strip()
    # "3/1(月)" のような曜日付きを除去
    for ch in "（(":
        if ch in s:
            s = s[: s.index(ch)].strip()
    # "3/1", "03/01", "2026/3/1" のようなパターン
    for fmt in ("%m/%d", "%Y/%m/%d", "%m月%d日"):
        try:
            datetime.strptime(s, fmt)
            return True
        except ValueError:
            pass
    # 整数 1〜31 は日付（日）とみなす
    try:
        n = int(s)
        if 1 <= n <= 31:
            return True
    except (ValueError, TypeError):
        pass
    return False


def detect_name_column(ws, header_row, max_scan=10):
    """従業員名の列を自動検出する。

    ヘッダー行より下に文字列が入っている最も左の列を名前列とみなす。
    """
    for col_idx in range(1, min(max_scan + 1, ws.max_column + 1)):
        text_count = 0
        for row_idx in range(header_row + 1, min(header_row + 11, ws.max_row + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
                text_count += 1
        if text_count >= 2:
            return col_idx
    return 1


def parse_dates(ws, header_row, name_col, year=None):
    """ヘッダー行から日付マッピング {列番号: 日付文字列} を構築する。"""
    if year is None:
        year = datetime.now().year

    dates = {}
    for col_idx in range(1, ws.max_column + 1):
        if col_idx == name_col:
            continue
        cell = ws.cell(row=header_row, column=col_idx)
        value = cell.value
        if value is None:
            continue

        date_str = _normalize_date(value, year)
        if date_str:
            dates[col_idx] = date_str
    return dates


def _normalize_date(value, year):
    """各種日付表現を YYYY-MM-DD 形式に正規化する。"""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    s = str(value).strip()

    # "3/1(月)" のような曜日付きを除去
    for ch in "（(":
        if ch in s:
            s = s[: s.index(ch)].strip()

    for fmt, has_year in [
        ("%Y/%m/%d", True),
        ("%Y-%m-%d", True),
        ("%m/%d", False),
        ("%m月%d日", False),
    ]:
        try:
            dt = datetime.strptime(s, fmt)
            if not has_year:
                dt = dt.replace(year=year)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            pass

    # 整数のみ（日にちのみ）の場合はスキップ（月が不明）
    return None


def find_data_rows(ws, header_row, name_col):
    """従業員データの行範囲を特定する。"""
    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=name_col)
        if cell.value and str(cell.value).strip():
            rows.append(row_idx)
    return rows


def convert_sheet(ws, year=None):
    """ワークシート1枚をパースし、シフトレコードのリストを返す。

    Returns:
        list[dict]: [{"date": "2026-03-01", "employee": "田中太郎", "shift": "早番"}, ...]
    """
    header_row = detect_header_row(ws)
    if header_row is None:
        raise ValueError(
            f"シート '{ws.title}' から日付ヘッダー行を検出できませんでした。"
        )

    name_col = detect_name_column(ws, header_row)
    dates = parse_dates(ws, header_row, name_col, year=year)
    if not dates:
        raise ValueError(
            f"シート '{ws.title}' から日付列を検出できませんでした。"
        )

    data_rows = find_data_rows(ws, header_row, name_col)
    records = []

    for row_idx in data_rows:
        employee = str(ws.cell(row=row_idx, column=name_col).value).strip()
        for col_idx, date_str in dates.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            shift_value = cell.value
            if shift_value is None:
                continue
            shift_str = str(shift_value).strip()
            if not shift_str:
                continue
            records.append({
                "date": date_str,
                "employee": employee,
                "shift": shift_str,
            })

    return records


def convert_workbook(excel_path, sheet_name=None, year=None):
    """Excelブック全体（または指定シート）を変換する。"""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    all_records = []

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"シート '{sheet_name}' が見つかりません。"
                f"利用可能: {wb.sheetnames}"
            )
        sheets = [wb[sheet_name]]
    else:
        sheets = list(wb.worksheets)

    for ws in sheets:
        try:
            records = convert_sheet(ws, year=year)
            all_records.extend(records)
        except ValueError as e:
            print(f"警告: {e} (スキップします)", file=sys.stderr)

    wb.close()
    return all_records


def write_csv(records, output_path, encoding="utf-8-sig"):
    """レコードをCSVファイルに書き出す。"""
    fieldnames = ["date", "employee", "shift"]

    with open(output_path, "w", newline="", encoding=encoding) as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        # 日付→従業員名でソート
        for record in sorted(records, key=lambda r: (r["date"], r["employee"])):
            writer.writerow(record)

    return len(records)


def main():
    parser = argparse.ArgumentParser(
        description="ExcelシフトデータをCSV形式に変換するツール",
    )
    parser.add_argument(
        "input",
        help="入力ファイルのパス (.xlsx または .csv)",
    )
    parser.add_argument(
        "-o", "--output",
        help="出力CSVファイルのパス (デフォルト: 入力ファイル名.csv)",
    )
    parser.add_argument(
        "-s", "--sheet",
        help="変換対象のシート名 (未指定で全シート)",
    )
    parser.add_argument(
        "-y", "--year",
        type=int,
        help="日付の年 (ヘッダーに年が無い場合に使用、デフォルト: 今年)",
    )
    parser.add_argument(
        "-e", "--encoding",
        default="utf-8-sig",
        help="出力CSVのエンコーディング (デフォルト: utf-8-sig)",
    )
    parser.add_argument(
        "-f", "--format",
        choices=["flat", "freee"],
        default="flat",
        help="出力形式: flat (デフォルト) または freee (freee勤怠管理Plus用)",
    )
    parser.add_argument(
        "-m", "--mapping",
        default="config/freee_mapping.json",
        help="freee形式用のマッピングファイル (デフォルト: config/freee_mapping.json)",
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"エラー: ファイルが見つかりません: {input_path}", file=sys.stderr)
        sys.exit(1)

    # 出力パス決定
    if args.output:
        output_path = Path(args.output)
    elif args.format == "freee":
        output_path = input_path.with_stem(input_path.stem + "_freee").with_suffix(".csv")
    elif input_path.suffix.lower() == ".csv":
        output_path = input_path.with_stem(input_path.stem + "_flat")
    else:
        output_path = input_path.with_suffix(".csv")

    print(f"入力: {input_path}")
    print(f"出力: {output_path}")
    print(f"形式: {args.format}")

    if args.format == "freee":
        records = convert_to_freee(input_path, args.mapping, encoding=args.encoding)
        if not records:
            print("警告: 変換できるシフトデータがありませんでした。", file=sys.stderr)
            print("マッピングファイルを確認してください:", args.mapping, file=sys.stderr)
            sys.exit(1)
        count = write_freee_csv(records, output_path)
    else:
        if input_path.suffix.lower() == ".csv":
            records = convert_csv_file(input_path, year=args.year, encoding=args.encoding)
        else:
            records = convert_workbook(input_path, sheet_name=args.sheet, year=args.year)

        if not records:
            print("警告: シフトデータが見つかりませんでした。", file=sys.stderr)
            sys.exit(1)

        count = write_csv(records, output_path, encoding=args.encoding)

    print(f"完了: {count}件のシフトレコードを出力しました。")


if __name__ == "__main__":
    main()
