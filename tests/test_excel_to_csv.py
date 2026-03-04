"""excel_to_csv モジュールのテスト。"""

import csv
import json
import tempfile
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.excel_to_csv import (
    CsvWorksheet,
    convert_csv_file,
    convert_sheet,
    convert_to_freee,
    convert_workbook,
    detect_header_row,
    detect_name_column,
    parse_dates,
    write_csv,
    write_freee_csv,
)

WEEKDAYS = ["月", "火", "水", "木", "金", "土", "日"]


def _create_shift_workbook(tmp_path, filename="test_shift.xlsx"):
    """テスト用のシフト表Excelを作成する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "3月シフト表"

    # タイトル行
    ws.cell(row=1, column=1, value="2026年3月 シフト表")

    # ヘッダー行 (3行目)
    ws.cell(row=3, column=1, value="従業員名")
    for day in range(1, 8):
        dt = datetime(2026, 3, day)
        weekday = WEEKDAYS[dt.weekday()]
        ws.cell(row=3, column=day + 1, value=f"3/{day}({weekday})")

    # 従業員データ
    employees = {
        "田中太郎": ["早番", "遅番", "夜勤", "休み", "早番", "休み", "休み"],
        "鈴木花子": ["遅番", "早番", "早番", "夜勤", "遅番", "休み", "休み"],
    }
    for i, (name, shifts) in enumerate(employees.items()):
        row = 4 + i
        ws.cell(row=row, column=1, value=name)
        for j, shift in enumerate(shifts):
            ws.cell(row=row, column=j + 2, value=shift)

    path = tmp_path / filename
    wb.save(path)
    return path


class TestDetectHeaderRow:
    def test_finds_header_with_date_strings(self, tmp_path):
        path = _create_shift_workbook(tmp_path)
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="タイトル")
        ws.cell(row=3, column=2, value="3/1(月)")
        ws.cell(row=3, column=3, value="3/2(火)")
        ws.cell(row=3, column=4, value="3/3(水)")

        assert detect_header_row(ws) == 3

    def test_returns_none_when_no_dates(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="テスト")
        ws.cell(row=2, column=1, value="データ")

        assert detect_header_row(ws) is None


class TestDetectNameColumn:
    def test_finds_name_column(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=3, column=1, value="従業員名")
        ws.cell(row=4, column=1, value="田中太郎")
        ws.cell(row=5, column=1, value="鈴木花子")

        assert detect_name_column(ws, header_row=3) == 1


class TestParseDates:
    def test_parses_slash_dates_with_weekday(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="従業員名")
        ws.cell(row=1, column=2, value="3/1(月)")
        ws.cell(row=1, column=3, value="3/2(火)")

        dates = parse_dates(ws, header_row=1, name_col=1, year=2026)
        assert dates[2] == "2026-03-01"
        assert dates[3] == "2026-03-02"

    def test_parses_datetime_objects(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="名前")
        ws.cell(row=1, column=2, value=datetime(2026, 3, 1))
        ws.cell(row=1, column=3, value=datetime(2026, 3, 2))

        dates = parse_dates(ws, header_row=1, name_col=1, year=2026)
        assert dates[2] == "2026-03-01"
        assert dates[3] == "2026-03-02"


class TestConvertSheet:
    def test_converts_sheet_to_records(self, tmp_path):
        path = _create_shift_workbook(tmp_path)
        from openpyxl import load_workbook

        wb = load_workbook(path)
        ws = wb.active

        records = convert_sheet(ws, year=2026)

        assert len(records) == 14  # 2 employees × 7 days
        # 特定レコードの確認
        tanaka_mar1 = [
            r for r in records
            if r["employee"] == "田中太郎" and r["date"] == "2026-03-01"
        ]
        assert len(tanaka_mar1) == 1
        assert tanaka_mar1[0]["shift"] == "早番"

    def test_skips_empty_cells(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="名前")
        ws.cell(row=1, column=2, value="3/1(月)")
        ws.cell(row=1, column=3, value="3/2(火)")
        ws.cell(row=2, column=1, value="田中太郎")
        ws.cell(row=2, column=2, value="早番")
        # 3/2 は空欄

        records = convert_sheet(ws, year=2026)
        assert len(records) == 1
        assert records[0]["shift"] == "早番"


class TestConvertWorkbook:
    def test_full_conversion(self, tmp_path):
        path = _create_shift_workbook(tmp_path)
        records = convert_workbook(path, year=2026)

        assert len(records) == 14
        employees = {r["employee"] for r in records}
        assert employees == {"田中太郎", "鈴木花子"}


class TestWriteCsv:
    def test_writes_sorted_csv(self, tmp_path):
        records = [
            {"date": "2026-03-02", "employee": "鈴木花子", "shift": "遅番"},
            {"date": "2026-03-01", "employee": "田中太郎", "shift": "早番"},
            {"date": "2026-03-01", "employee": "鈴木花子", "shift": "早番"},
        ]
        output_path = tmp_path / "output.csv"
        count = write_csv(records, output_path)

        assert count == 3
        with open(output_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 3
        # 日付→名前の順でソートされている
        assert rows[0]["date"] == "2026-03-01"
        assert rows[0]["employee"] == "田中太郎"
        assert rows[1]["date"] == "2026-03-01"
        assert rows[1]["employee"] == "鈴木花子"
        assert rows[2]["date"] == "2026-03-02"


def _create_shift_csv(tmp_path, filename="test_shift.csv"):
    """テスト用のシフト表CSVを作成する。"""
    path = tmp_path / filename
    rows = [
        ["2026年3月 シフト表"],
        [],
        ["従業員名", "3/1(月)", "3/2(火)", "3/3(水)", "3/4(木)", "3/5(金)", "3/6(土)", "3/7(日)"],
        ["田中太郎", "早番", "遅番", "夜勤", "休み", "早番", "休み", "休み"],
        ["鈴木花子", "遅番", "早番", "早番", "夜勤", "遅番", "休み", "休み"],
    ]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    return path


class TestCsvWorksheet:
    def test_from_file(self, tmp_path):
        path = _create_shift_csv(tmp_path)
        ws = CsvWorksheet.from_file(path)

        assert ws.max_row == 5
        assert ws.cell(row=3, column=1).value == "従業員名"
        assert ws.cell(row=3, column=2).value == "3/1(月)"
        assert ws.cell(row=4, column=1).value == "田中太郎"

    def test_detect_header_row_on_csv(self, tmp_path):
        path = _create_shift_csv(tmp_path)
        ws = CsvWorksheet.from_file(path)

        assert detect_header_row(ws) == 3

    def test_convert_sheet_on_csv(self, tmp_path):
        path = _create_shift_csv(tmp_path)
        ws = CsvWorksheet.from_file(path)

        records = convert_sheet(ws, year=2026)
        assert len(records) == 14
        tanaka = [r for r in records if r["employee"] == "田中太郎" and r["date"] == "2026-03-01"]
        assert len(tanaka) == 1
        assert tanaka[0]["shift"] == "早番"


class TestConvertCsvFile:
    def test_csv_input_full_conversion(self, tmp_path):
        path = _create_shift_csv(tmp_path)
        records = convert_csv_file(path, year=2026)

        assert len(records) == 14
        employees = {r["employee"] for r in records}
        assert employees == {"田中太郎", "鈴木花子"}

    def test_csv_to_csv_roundtrip(self, tmp_path):
        """CSV入力 → 変換 → フラットCSV出力の一連の流れをテスト。"""
        csv_input = _create_shift_csv(tmp_path)
        csv_output = tmp_path / "output.csv"

        records = convert_csv_file(csv_input, year=2026)
        write_csv(records, csv_output)

        with open(csv_output, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 14
        assert all(r["date"].startswith("2026-03-") for r in rows)
        assert set(r["employee"] for r in rows) == {"田中太郎", "鈴木花子"}


def _create_tabular_csv(tmp_path, filename="test_tabular.csv"):
    """テスト用のテーブル形式シフトCSVを作成する（1行1レコード）。"""
    path = tmp_path / filename
    rows = [
        ["指導員名", "職種", "日付", "勤務時間(開始)", "勤務時間(終了)", "シフト名"],
        ["梶井大成", "管理者児童指導員", "2026/4/1", "8:00", "17:00", "早出出勤（8-17）"],
        ["梶井大成", "管理者児童指導員", "2026/4/2", "9:00", "18:00", "通常出勤（9-18）"],
        ["梶井大成", "管理者児童指導員", "2026/4/3", "", "", "公休"],
        ["山田花子", "児童指導員", "2026/4/1", "9:00", "18:00", "通常出勤（9-18）"],
        ["山田花子", "児童指導員", "2026/4/2", "10:00", "19:00", "遅出出勤（10-19）"],
    ]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    return path


class TestTabularCsv:
    def test_detects_tabular_format(self, tmp_path):
        path = _create_tabular_csv(tmp_path)
        records = convert_csv_file(path)

        assert len(records) == 5
        employees = {r["employee"] for r in records}
        assert employees == {"梶井大成", "山田花子"}

    def test_tabular_date_normalized(self, tmp_path):
        path = _create_tabular_csv(tmp_path)
        records = convert_csv_file(path)

        dates = {r["date"] for r in records}
        assert "2026-04-01" in dates
        assert "2026-04-02" in dates

    def test_tabular_shift_names_preserved(self, tmp_path):
        path = _create_tabular_csv(tmp_path)
        records = convert_csv_file(path)

        shifts = {r["shift"] for r in records}
        assert "早出出勤（8-17）" in shifts
        assert "公休" in shifts

    def test_tabular_csv_to_flat_csv(self, tmp_path):
        """テーブル形式CSV → フラットCSV出力の流れをテスト。"""
        csv_input = _create_tabular_csv(tmp_path)
        csv_output = tmp_path / "output.csv"

        records = convert_csv_file(csv_input)
        write_csv(records, csv_output)

        with open(csv_output, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 5
        assert set(rows[0].keys()) == {"date", "employee", "shift"}


def _create_freee_mapping(tmp_path, filename="freee_mapping.json"):
    """テスト用のfreeeマッピングファイルを作成する。"""
    path = tmp_path / filename
    mapping = {
        "employee_map": {
            "梶井大成": "107",
            "山田花子": "108",
        },
        "time_to_pattern": {
            "8:00-17:00": "早出",
            "9:00-18:00": "特別1",
            "10:00-19:00": "通常",
        },
        "shift_map": {},
        "skip_shifts": ["公休"],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False)
    return path


class TestFreeeConversion:
    def test_converts_to_freee_format(self, tmp_path):
        csv_path = _create_tabular_csv(tmp_path)
        mapping_path = _create_freee_mapping(tmp_path)

        records = convert_to_freee(csv_path, mapping_path)

        assert len(records) == 4  # 5 rows - 1 公休 = 4
        assert all("勤務日" in r for r in records)
        assert all("従業員コード" in r for r in records)
        assert all("パターンコード" in r for r in records)

    def test_employee_code_mapping(self, tmp_path):
        csv_path = _create_tabular_csv(tmp_path)
        mapping_path = _create_freee_mapping(tmp_path)

        records = convert_to_freee(csv_path, mapping_path)

        codes = {r["従業員コード"] for r in records}
        assert codes == {"107", "108"}

    def test_time_based_pattern_matching(self, tmp_path):
        csv_path = _create_tabular_csv(tmp_path)
        mapping_path = _create_freee_mapping(tmp_path)

        records = convert_to_freee(csv_path, mapping_path)

        patterns = {r["パターンコード"] for r in records}
        assert "早出" in patterns
        assert "特別1" in patterns
        assert "通常" in patterns

    def test_skips_public_holidays(self, tmp_path):
        csv_path = _create_tabular_csv(tmp_path)
        mapping_path = _create_freee_mapping(tmp_path)

        records = convert_to_freee(csv_path, mapping_path)

        # 公休はスキップされる
        shifts = {r["パターンコード"] for r in records}
        assert "公休" not in shifts

    def test_date_format_uses_slash(self, tmp_path):
        csv_path = _create_tabular_csv(tmp_path)
        mapping_path = _create_freee_mapping(tmp_path)

        records = convert_to_freee(csv_path, mapping_path)

        for r in records:
            assert "/" in r["勤務日"]
            assert "-" not in r["勤務日"]

    def test_write_freee_csv(self, tmp_path):
        csv_path = _create_tabular_csv(tmp_path)
        mapping_path = _create_freee_mapping(tmp_path)
        output_path = tmp_path / "freee_output.csv"

        records = convert_to_freee(csv_path, mapping_path)
        count = write_freee_csv(records, output_path)

        assert count == 4
        with open(output_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 4
        assert set(rows[0].keys()) == {"勤務日", "従業員コード", "パターンコード"}

    def test_space_in_employee_name(self, tmp_path):
        """スペース入りの名前でもマッチする。"""
        csv_input = tmp_path / "spaced.csv"
        rows = [
            ["指導員名", "職種", "日付", "勤務時間(開始)", "勤務時間(終了)", "シフト名"],
            ["梶井 大成", "管理者", "2026/4/1", "8:00", "17:00", "早出出勤（8-17）"],
        ]
        with open(csv_input, "w", newline="", encoding="utf-8-sig") as f:
            csv.writer(f).writerows(rows)

        mapping_path = _create_freee_mapping(tmp_path)
        records = convert_to_freee(csv_input, mapping_path)

        assert len(records) == 1
        assert records[0]["従業員コード"] == "107"


class TestEndToEnd:
    def test_excel_to_csv_roundtrip(self, tmp_path):
        """Excel生成 → 変換 → CSV読み込みの一連の流れをテスト。"""
        excel_path = _create_shift_workbook(tmp_path)
        csv_path = tmp_path / "output.csv"

        records = convert_workbook(excel_path, year=2026)
        write_csv(records, csv_path)

        with open(csv_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 14
        assert all(r["date"].startswith("2026-03-") for r in rows)
        assert set(r["employee"] for r in rows) == {"田中太郎", "鈴木花子"}
        shift_types = set(r["shift"] for r in rows)
        assert shift_types <= {"早番", "遅番", "夜勤", "休み"}
