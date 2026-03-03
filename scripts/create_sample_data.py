"""サンプルのシフト表Excelファイルを生成するスクリプト。"""

import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SHIFT_TYPES = ["早番", "遅番", "夜勤", "休み", "有休"]
EMPLOYEES = ["田中太郎", "鈴木花子", "佐藤一郎", "高橋美咲", "伊藤健太"]
WEEKDAYS = ["月", "火", "水", "木", "金", "土", "日"]


def create_sample_excel(output_path, year=2026, month=3):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}月シフト表"

    # タイトル行
    ws.cell(row=1, column=1, value=f"{year}年{month}月 シフト表")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # 日付ヘッダー (3行目)
    header_row = 3
    ws.cell(row=header_row, column=1, value="従業員名")
    ws.cell(row=header_row, column=1).font = Font(bold=True)

    # 月の日数を計算
    if month == 12:
        days_in_month = (datetime(year, 1, 1) - datetime(year, 12, 1)).days
    else:
        days_in_month = (datetime(year, month + 1, 1) - datetime(year, month, 1)).days

    for day in range(1, days_in_month + 1):
        col = day + 1
        dt = datetime(year, month, day)
        weekday = WEEKDAYS[dt.weekday()]
        ws.cell(row=header_row, column=col, value=f"{month}/{day}({weekday})")
        ws.cell(row=header_row, column=col).font = Font(bold=True)
        ws.cell(row=header_row, column=col).alignment = Alignment(horizontal="center")

        # 土日は背景色を変える
        if dt.weekday() >= 5:
            ws.cell(row=header_row, column=col).fill = PatternFill(
                start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"
            )

    # 従業員データ (4行目〜)
    random.seed(42)
    for i, name in enumerate(EMPLOYEES):
        row = header_row + 1 + i
        ws.cell(row=row, column=1, value=name)

        for day in range(1, days_in_month + 1):
            col = day + 1
            dt = datetime(year, month, day)

            # 土日は休みになりやすい
            if dt.weekday() >= 5:
                shift = random.choice(["休み", "休み", "休み", "早番", "夜勤"])
            else:
                shift = random.choice(SHIFT_TYPES[:3] + ["早番"])  # 平日は勤務が多め
            ws.cell(row=row, column=col, value=shift)
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

    # 列幅の調整
    ws.column_dimensions["A"].width = 14
    for day in range(1, days_in_month + 1):
        col_letter = chr(ord("A") + day)
        if day + 1 <= 26:
            ws.column_dimensions[chr(ord("A") + day)].width = 12

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"サンプルExcelファイルを作成しました: {output_path}")


if __name__ == "__main__":
    create_sample_excel("sample_data/shift_sample.xlsx")
